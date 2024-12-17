from django.shortcuts import render
from openpyxl import load_workbook
from .forms import SearchForm

# Create your views here.

def search_employee(request):
    form = SearchForm()
    items = []
    if request.method == 'POST':
        form = SearchForm(request.POST)
        if form.is_valid():
            cd = form.cleaned_data
            wb = load_workbook(r'C:\Users\East-Sound\Desktop\Django Projects\Video_12\Employees.xlsx')
            ws = wb['emp']
            rows_num = ws.max_row
            for i in range(1, rows_num + 1):
                if cd['name'].upper() == ws.cell(row=i, column=1).value.upper():
                    item = []
                    item.append(ws.cell(row=i, column=1).value)
                    item.append(ws.cell(row=i, column=2).value)
                    item.append(ws.cell(row=i, column=3).value)
                    item.append(ws.cell(row=i, column=4).value)
                    items.append(item)
    return render(request, 'pages/display_search.html', {'items': items, 'form': form})
