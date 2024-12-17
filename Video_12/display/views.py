from django.shortcuts import render
from openpyxl import load_workbook

# Create your views here.

def display_employee(request):
    wb = load_workbook(r'C:\Users\East-Sound\Desktop\Django Projects\Video_12\Employees.xlsx')
    ws = wb['emp']
    rows_num = ws.max_row
    items = []
    for i in range(1, rows_num + 1):
        item = []
        item.append(ws.cell(row=i, column=1).value)
        item.append(ws.cell(row=i, column=2).value)
        item.append(ws.cell(row=i, column=3).value)
        item.append(ws.cell(row=i, column=4).value)
        items.append(item)
    return render(request, 'pages/display_emp.html', {'items': items})