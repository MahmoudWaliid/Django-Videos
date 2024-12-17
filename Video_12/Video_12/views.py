from django.shortcuts import render
from .forms import RegisterForm
from openpyxl import load_workbook

def new_employee(request):
    form = RegisterForm()
    o = ''
    if request.method == 'POST':
        form = RegisterForm(request.POST)
        if form.is_valid():
            wb = load_workbook(r'C:\Users\East-Sound\Desktop\Django Projects\Video_12\Employees.xlsx')
            ws = wb['emp']
            
            current_row = ws.max_row + 1
            
            ws[f'A{current_row}'] = form.cleaned_data['name']
            ws[f'B{current_row}'] = form.cleaned_data['salary']
            ws[f'C{current_row}'] = form.cleaned_data['email']
            ws[f'D{current_row}'] = form.cleaned_data['address']
            
            wb.save(r'C:\Users\East-Sound\Desktop\Django Projects\Video_12\Employees.xlsx')
            
            o = 'Employee added successfully!'
            form = RegisterForm()
            
    return render(request, 'pages/new_emp.html', {'form': form, 'output': o})
